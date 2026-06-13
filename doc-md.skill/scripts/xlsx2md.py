#!/usr/bin/env python3
"""Convert Excel file to Markdown tables."""

import sys
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def xlsx_to_markdown(input_path: str, output_path: str = None) -> str:
    """Convert xlsx file to markdown."""
    input_file = Path(input_path)
    if not input_file.exists():
        raise FileNotFoundError(f"File not found: {input_path}")
    
    if output_path is None:
        output_path = input_file.with_suffix('.md')
    
    wb = openpyxl.load_workbook(input_path, data_only=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        for sheet_name in wb.sheetnames:
            f.write(f"# Sheet: {sheet_name}\n\n")
            
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            
            if not rows:
                f.write("*(Empty sheet)*\n\n")
                continue
            
            # Write header
            header = rows[0]
            f.write('| ' + ' | '.join(str(cell) if cell is not None else '' for cell in header) + ' |\n')
            f.write('|' + '|'.join('---' for _ in header) + '|\n')
            
            # Write data rows
            for row in rows[1:]:
                f.write('| ' + ' | '.join(str(cell) if cell is not None else '' for cell in row) + ' |\n')
            
            f.write('\n')
    
    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python xlsx2md.py <input.xlsx> [output.md]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        result = xlsx_to_markdown(input_file, output_file)
        print(f"Converted: {result}")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
